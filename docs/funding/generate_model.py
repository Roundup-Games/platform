#!/usr/bin/env python3
"""
Generator for the Roundup Games non-profit financial model (.xlsx).

Produces an editable, formula-driven workbook:
  - Read Me
  - Headcount     (role x dept x FTE x all-in cost -> drives the Personnel line)
  - Assumptions   (all non-personnel drivers; yellow input cells)
  - Model - Conservative / Base / Ambitious  (formula-driven from Headcount + Assumptions)
  - Research & Benchmarks
  - Summary

Run with the build-tools venv python (has openpyxl installed):
  build-tools/.venv/bin/python docs/funding/generate_model.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# HEADCOUNT: roles, departments, all-in annual cost per 1.0 FTE (EUR, employer cost).
# All-in = gross salary x ~1.22 (Arbeitgeberanteil ~21-23%). See Research sheet.
# --------------------------------------------------------------------------- #
# (key, label, department, cost_eur_per_fte)
ROLES = [
    ("lead_eng",    "Lead / Senior Engineer",      "Engineering",                90000),
    ("mid_eng",     "Mid Engineer",                "Engineering",                65000),
    ("community",   "Community Manager",           "Community & Operations",     50000),
    ("hq_coord",    "HQ / Space Coordinator",      "Community & Operations",     42000),
    ("events",      "Event Coordinator",           "Community & Operations",     45000),
    ("fundraising", "Fundraising / Grants Lead",   "Partnerships & Fundraising", 60000),
    ("partners",    "Partnerships Manager",        "Partnerships & Fundraising", 60000),
    ("library",     "Library Coordinator",         "Programs",                   40000),
    ("cert",        "Certification Lead",          "Programs",                   55000),
    ("finance",     "Finance / Office Manager",    "Finance & Admin",            45000),
]

# FTE per scenario per year [Y1..Y5]. Fractional = part-time. 0 = not staffed (volunteer/outsourced).
FTE = {
    "Conservative": {
        "lead_eng":    [0,   0,   0,    0.05, 0.1 ],
        "mid_eng":     [0,   0,   0,    0,    0   ],
        "community":   [0,   0,   0.05, 0.1,  0.15],
        "hq_coord":    [0,   0,   0,    0,    0.15],
        "events":      [0,   0,   0,    0,    0   ],
        "fundraising": [0.1, 0.1, 0.15, 0.2,  0.25],
        "partners":    [0,   0,   0,    0,    0   ],
        "library":     [0,   0,   0,    0,    0   ],
        "cert":        [0,   0,   0,    0,    0   ],
        "finance":     [0,   0,   0,    0,    0   ],  # outsourced via Legal line until scale
    },
    "Base": {
        "lead_eng":    [0.5, 1.0, 1.0,  1.0,  1.0 ],   # LOCKED (engineering)
        "mid_eng":     [0,   0.5, 0.5,  0.5,  0.5 ],   # LOCKED (engineering)
        "community":   [0.3, 0.4, 0.5,  0.7,  0.9 ],
        "hq_coord":    [0.2, 0.3, 0.4,  0.5,  0.7 ],
        "events":      [0,   0.1, 0.2,  0.3,  0.4 ],
        "fundraising": [0.15,0.25,0.3,  0.3,  0.4 ],
        "partners":    [0,   0.1, 0.25, 0.35, 0.4 ],
        "library":     [0,   0.1, 0.15, 0.2,  0.3 ],
        "cert":        [0,   0.1, 0.2,  0.3,  0.4 ],
        "finance":     [0,   0,   0,    0.1,  0.2 ],
    },
    "Ambitious": {
        "lead_eng":    [0.5, 1.0, 1.0,  1.5,  2.0 ],
        "mid_eng":     [0,   0.5, 0.5,  1.0,  1.0 ],
        "community":   [0.5, 0.5, 1.0,  1.0,  1.0 ],
        "hq_coord":    [0.2, 0.5, 0.5,  1.0,  1.0 ],
        "events":      [0,   0.1, 0.2,  0.5,  0.5 ],
        "fundraising": [0.15,0.25,0.35, 0.5,  0.5 ],
        "partners":    [0,   0.2, 0.2,  0.4,  0.4 ],
        "library":     [0,   0.1, 0.2,  0.25, 0.4 ],
        "cert":        [0,   0.2, 0.4,  0.5,  0.7 ],
        "finance":     [0,   0,   0,    0.1,  0.2 ],
    },
}

# Headcount sheet rows
HC_ROLE_START_ROW = 5
HC_COST_ROW = HC_ROLE_START_ROW + len(ROLES)        # SUMPRODUCT personnel cost
HC_FTE_ROW  = HC_COST_ROW + 1                       # SUM total FTE

# --------------------------------------------------------------------------- #
# Non-personnel assumption rows.
# --------------------------------------------------------------------------- #
ASSUMPTION_ROWS = [
    ("mau",        "Monthly active users",          "users",   [300,700,1200,2000,3000],   [1200,3000,8000,21000,50000], [3000,8000,25000,60000,120000],   "#,##0"),
    ("paid_seats", "Paid seats sold",               "seats/yr",[150,400,900,1600,2500],   [800,2500,6500,14000,27000],   [2500,7000,20000,38000,75000],   "#,##0"),
    ("seat_price", "Average seat price",            "EUR",     [16,16,17,17,18],           [18,19,20,21,22],             [18,19,21,22,24],               "0.00 \"EUR\""),
    ("comm_rate",  "Commission take rate",          "%",       [0.06,0.06,0.07,0.07,0.08], [0.09,0.10,0.11,0.12,0.13],   [0.08,0.09,0.10,0.10,0.11],     "0.0%"),
    ("gm_count",   "Paying GMs (avg in year)",      "GMs",     [5,12,22,35,55],            [15,45,90,160,260],           [40,110,230,420,700],           "#,##0"),
    ("gm_price",   "GM subscription price",         "EUR/mo",  [7.99,7.99,8.99,8.99,9.99], [10.99,11.99,12.99,13.99,14.99],[9.99,10.99,11.99,12.99,13.99], "0.00 \"EUR\""),
    ("events",     "Branded events",                "events",  [0,1,1,2,2],                [1,2,4,8,16],                  [2,4,8,16,32],                  "#,##0"),
    ("att",        "Avg attendance per event",      "people",  [0,50,70,100,120],          [80,120,180,250,350],         [100,180,300,450,600],          "#,##0"),
    ("evt_price",  "Event ticket price",            "EUR",     [0,12,14,15,16],            [15,18,22,25,28],             [16,20,25,28,32],               "0.00 \"EUR\""),
    ("evt_cogs",   "Event cost per attendee",       "EUR",     [0,9,10,10,10],             [11,12,14,15,16],             [11,12,14,15,16],               "0.00 \"EUR\""),
    ("lot_members","Library-tier members",          "members", [0,20,35,55,80],            [30,80,150,250,380],          [60,160,320,550,850],           "#,##0"),
    ("lot_price",  "Library tier fee",              "EUR/mo",  [3.99,3.99,4.49,4.49,4.99], [4.99,5.49,5.99,6.49,6.99],   [5.99,6.49,6.99,7.49,7.99],     "0.00 \"EUR\""),
    ("lot_ops",    "Library shrinkage & replenish", "EUR/yr",  [0,300,350,400,600],        [400,450,500,700,800],        [500,600,900,1300,1500],        "#,##0 \"EUR\""),
    ("lot_inv",    "Library inventory (partnership)","EUR/yr", [0,1500,0,0,1500],          [2000,0,0,2000,0],            [2500,0,2500,2500,0],           "#,##0 \"EUR\""),
    ("certs",      "New certifications",            "certs/yr",[0,8,12,18,25],             [15,30,50,75,110],            [25,60,110,180,260],            "#,##0"),
    ("cert_fee",   "Certification fee",             "EUR",     [0,99,99,129,129],          [179,199,229,259,289],        [179,199,219,249,269],          "0.00 \"EUR\""),
    ("cert_dev",   "Curriculum development",        "EUR/yr",  [500,200,200,300,300],      [3000,500,500,1000,1000],     [5000,1000,1500,2500,2500],     "#,##0 \"EUR\""),
    ("cert_ops",   "Assessment + credentialing",    "EUR/cert",[0,25,25,30,30],            [30,30,35,40,45],             [35,40,45,50,55],               "0.00 \"EUR\""),
    ("b2b_events", "B2B events booked",             "events",  [0,2,4,7,10],               [6,16,36,75,150],               [6,18,40,100,200],               "#,##0"),
    ("b2b_att",    "Attendance per B2B event",      "people",  [0,15,18,20,22],            [18,20,22,25,28],             [20,22,25,28,30],               "#,##0"),
    ("b2b_price",  "B2B price per attendee",        "EUR",     [0,35,38,40,42],            [42,48,55,62,70],             [45,50,55,60,65],               "0.00 \"EUR\""),
    ("b2b_comm",   "B2B commission rate",           "%",       [0.10,0.10,0.12,0.12,0.15], [0.13,0.15,0.17,0.18,0.20],   [0.13,0.15,0.16,0.18,0.20],     "0.0%"),
    ("grants",     "Grants (project + program)",    "EUR/yr",  [15000,25000,40000,50000,60000],[55000,115000,200000,180000,160000],[80000,160000,260000,350000,420000], "#,##0 \"EUR\""),
    ("donors",     "Recurring donors",              "donors",  [10,25,45,70,100],          [40,80,150,250,400],          [80,200,450,800,1300],          "#,##0"),
    ("don_avg",    "Avg recurring donation",        "EUR/yr",  [45,50,55,60,65],           [60,75,90,100,110],           [70,85,100,115,125],            "#,##0 \"EUR\""),
    ("don_oneoff", "One-off donations",             "EUR/yr",  [1500,2500,4000,5000,7000], [3000,5000,8000,12000,18000], [6000,12000,20000,32000,45000], "#,##0 \"EUR\""),
    ("bridge",     "Working-capital bridge (capacity grant / founder loan)", "EUR/yr", [0,0,0,0,0], [50000,80000,10000,30000,0], [45000,70000,0,0,0], "#,##0 \"EUR\""),
    ("locations",  "Community HQ locations",        "count",   [0,1,1,1,2],                [1,1,2,3,5],                  [1,2,3,5,8],                    "#,##0"),
    ("space_cost", "HQ cost (rent+utils+insurance)","EUR/mo",  [0,700,750,800,850],        [1000,1050,1100,1200,1300],   [1400,1500,1600,1700,1800],     "#,##0 \"EUR\""),
    ("space_setup","HQ fit-out / expansion (capex)","EUR/yr",  [0,6000,0,0,6000],          [8000,1000,0,10000,0],        [10000,2000,12000,12000,2000],  "#,##0 \"EUR\""),
    ("infra",      "Infrastructure & hosting",      "EUR/mo",  [300,400,600,800,1000],     [650,1000,1500,2000,2500],    [1200,2500,4500,7000,10000],    "#,##0 \"EUR\""),
    ("legal",      "Legal / admin / accounting",    "EUR/yr",  [2000,2500,3000,4000,5000], [3000,4000,6000,8000,10000],  [5000,8000,12000,18000,25000],  "#,##0 \"EUR\""),
    ("marketing",  "Marketing & community",         "EUR/yr",  [1000,2000,3000,5000,7000], [3000,6000,10000,15000,20000],[8000,18000,35000,60000,90000], "#,##0 \"EUR\""),
    ("pay_rate",   "Payment processing rate",       "%",       [0.035]*5,                  [0.035,0.035,0.035,0.032,0.030],[0.035,0.032,0.030,0.028,0.025],"0.0%"),
]

ASSUMPTION_START_ROW = 4
ROW_OF = {key: ASSUMPTION_START_ROW + i for i, (key, *_rest) in enumerate(ASSUMPTION_ROWS)}

SCENARIO_COLS = {
    "Conservative": ["D", "E", "F", "G", "H"],
    "Base":         ["J", "K", "L", "M", "N"],
    "Ambitious":    ["P", "Q", "R", "S", "T"],
}
MODEL_COLS = ["B", "C", "D", "E", "F"]

# --------------------------------------------------------------------------- #
# Styles
# --------------------------------------------------------------------------- #
TITLE   = Font(name="Inter", size=16, bold=True, color="1F2937")
H2      = Font(name="Inter", size=12, bold=True, color="835500")
BOLD    = Font(name="Inter", size=11, bold=True)
NORMAL  = Font(name="Inter", size=11)
SMALL   = Font(name="Inter", size=10, color="4B5563")
INPUT   = PatternFill("solid", fgColor="FEF3C7")
SECT    = PatternFill("solid", fgColor="F3E8D2")
TOTAL   = PatternFill("solid", fgColor="EDE4D3")
MEMO    = PatternFill("solid", fgColor="F1F5F9")
GROUP_FILL = PatternFill("solid", fgColor="F3E8D2")
WRAP    = Alignment(wrap_text=True, vertical="top")
RIGHT   = Alignment(horizontal="right")
EUR = '#,##0 "EUR"'
EUR2 = '#,##0.00 "EUR"'

wb = Workbook()

# =============================== READ ME ==================================== #
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 110
ws["B2"] = "Roundup Games — Non-Profit Financial Model"; ws["B2"].font = TITLE
ws["B3"] = "5-year projection · three scenarios · formula-driven · editable · headcount-driven personnel"
ws["B3"].font = SMALL

readme = [
    ("", ""),
    ("Purpose", "A discussion tool for collaborator / funder conversations about the viability of the "
                "non-profit organisation behind roundup.games."),
    ("How to use", "1) 'Headcount' sheet — edit FTE per role per year (yellow). Personnel cost is derived live "
                   "via SUMPRODUCT(role cost × FTE) and feeds every model.\n"
                   "2) 'Assumptions' sheet — all other drivers (growth, prices, grants, space).\n"
                   "3) 'Model - Conservative / Base / Ambitious' — full 5-year P&L per scenario.\n"
                   "4) 'Summary' compares scenarios; 'Research & Benchmarks' sources every figure."),
    ("Headcount model", "Personnel cost is NOT a flat number — it is built from a role × department × FTE matrix "
                        "with DACH all-in salary costs (gross × ~1.22 employer overhead). 10 roles across 5 "
                        "departments. Fractional FTE = part-time. Volunteer-leveraged: paid staff cover only "
                        "accountability/continuity roles; events, library counter, moderation, translation are "
                        "volunteer-staffed."),
    ("Departments", "Engineering · Community & Operations · Partnerships & Fundraising · Programs (Library + "
                    "Certification) · Finance & Admin. Finance/Admin stays OUTSOURCED (in the Legal line) until "
                    "Base Y4 / Ambitious Y3, when an in-house hire becomes cheaper than external bookkeeping."),
    ("The revenue streams", "1) Commission on paid seats · 2) GM subscriptions · 3) Branded events · "
                            "4) Library of things · 5) Certification · 6) B2B events · 7) Grants & donations."),
    ("Shared space", "Library + certification + events + B2B SHARE one community HQ per city. Space is modelled "
                     "ONCE (rent + utilities + fit-out), not per initiative. Locations scale 0→3."),
    ("Library sourcing", "Inventory via publisher/merchant/player partnerships (demo copies, discounts, donations), "
                         "NOT retail. Lending platform built IN-HOUSE (dev cost in Personnel; no licence fee)."),
    ("Key finding (honest)", "Realistic DACH headcount runs personnel at ~50–62% of revenue in early years. Base and "
                             "Ambitious therefore need a WORKING-CAPITAL BRIDGE (capacity grants + founder loan) "
                             "drawn across Y1–Y4 — shown as its own revenue line on each model. The bridge is the "
                             "explicit fundraising target: €170k for Base, €115k for Ambitious, €0 for Conservative. "
                             "With the bridge, every year is net-positive; the underlying operating business reaches "
                             "break-even at Y3 (Base) and turns genuinely surplus by Y5."),
    ("Working-capital bridge", "A dedicated revenue line — distinct from operating/project grants — that covers the "
                                "cumulative funding gap while commercial revenue scales. It is capacity-building grants "
                                "(non-repayable) plus a founder loan (repayable from Y5+ surplus). Drawn Y1–Y4, zero by "
                                "Y5 when the org is self-sustaining. Conservative needs none."),
    ("Tax note", "gemeinnützige Körperschaft (e.V./gGmbH). Purpose-serving revenue tax-exempt; commercial surplus "
                 "exempt to €45,000/yr then ~15% KSt (illustrative). Library + certification may qualify as "
                 "Zweckbetrieb (fully exempt) — an upside not modelled. Confirm with a Steuerberater."),
    ("Currency", "All figures in EUR, net of VAT (reverse-charge / small-business assumptions apply)."),

    ("Disclaimer", "Planning estimates for discussion only — not audited, not a funding guarantee."),
]
r = 5
for label, text in readme:
    ws.cell(row=r, column=2, value=label).font = H2
    r += 1
    c = ws.cell(row=r, column=2, value=text); c.font = NORMAL; c.alignment = WRAP
    ws.row_dimensions[r].height = 15 * (1 + text.count("\n") + len(text) // 95)
    r += 2

# =============================== HEADCOUNT ================================== #
def build_headcount():
    wh = wb.create_sheet("Headcount")
    wh.sheet_view.showGridLines = False
    wh["A1"] = "Headcount — role × department × FTE × all-in cost (drives Personnel)"; wh["A1"].font = TITLE
    wh["A2"] = ("Personnel cost = SUMPRODUCT(cost × FTE) per scenario-year. Edit yellow FTE cells. "
                "Costs are DACH all-in (gross × ~1.22 Arbeitgeberanteil)."); wh["A2"].font = SMALL
    wh.column_dimensions["A"].width = 30
    wh.column_dimensions["B"].width = 28
    wh.column_dimensions["C"].width = 16
    col = 4
    for scen in ["Conservative","Base","Ambitious"]:
        wh.cell(row=2, column=col, value=scen).font = H2
        wh.cell(row=2, column=col).alignment = Alignment(horizontal="center")
        wh.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+4)
        for yc in ["Y1","Y2","Y3","Y4","Y5"]:
            c = wh.cell(row=3, column=col, value=yc); c.font = BOLD; c.fill = SECT; c.alignment = Alignment(horizontal="center")
            col += 1
        col += 1
    # header labels
    for col_idx, lab in [(1,"Role"),(2,"Department"),(3,"€/yr per 1.0 FTE")]:
        c = wh.cell(row=3, column=col_idx, value=lab); c.font = BOLD; c.fill = SECT

    # roles
    for i, (key, label, dept, cost) in enumerate(ROLES):
        row = HC_ROLE_START_ROW + i
        wh.cell(row=row, column=1, value=label).font = NORMAL
        wh.cell(row=row, column=2, value=dept).font = SMALL
        cc = wh.cell(row=row, column=3, value=cost); cc.number_format = EUR; cc.font = NORMAL; cc.alignment = RIGHT
        col = 4
        for scen in ["Conservative","Base","Ambitious"]:
            for v in FTE[scen][key]:
                c = wh.cell(row=row, column=col, value=v)
                c.number_format = "0.00"; c.fill = INPUT; c.alignment = RIGHT; c.font = NORMAL
                col += 1
            col += 1

    # totals: personnel cost (SUMPRODUCT) + total FTE (SUM)
    cost_row = HC_COST_ROW
    fte_row = HC_FTE_ROW
    wh.cell(row=cost_row, column=1, value="Total personnel cost (EUR)").font = BOLD
    wh.cell(row=fte_row,  column=1, value="Total FTE").font = BOLD
    first = HC_ROLE_START_ROW
    last = HC_ROLE_START_ROW + len(ROLES) - 1
    col = 4
    for scen in ["Conservative","Base","Ambitious"]:
        for _ in range(5):
            L = get_column_letter(col)
            c = wh.cell(row=cost_row, column=col, value=f"=SUMPRODUCT($C${first}:$C${last},{L}${first}:{L}${last})")
            c.number_format = EUR; c.font = BOLD; c.fill = TOTAL; c.alignment = RIGHT
            f = wh.cell(row=fte_row, column=col, value=f"=SUM({L}{first}:{L}{last})")
            f.number_format = "0.00"; f.font = BOLD; f.fill = TOTAL; f.alignment = RIGHT
            col += 1
        col += 1

    # volunteer-leverage notes
    nr = fte_row + 2
    wh.cell(row=nr, column=1, value="Volunteer leverage (not in headcount)").font = H2
    notes = [
        "• Events: on-site setup/teardown, registration desk, game librarians at branded events — volunteer-staffed, "
        "overseen by the paid Event Coordinator (where present) or the Community Manager.",
        "• Library counter: lending/check-in at the HQ is volunteer-staffed (2–4 regulars per location), coordinated "
        "by the HQ/Space Coordinator. This is how Libraries of Things stay viable per the research.",
        "• Content: translation (EN/DE), community moderation backstop, documentation — volunteer-contributed.",
        "• Engineering: the platform is built; non-critical features and bug-fixes can take open-source-style "
        "volunteer contributions, reviewed by the paid Lead Engineer.",
        "• The paid roles above cover ONLY accountability/continuity: fundraising, finance/compliance, platform "
        "integrity, HQ operations, and (at scale) programme delivery.",
    ]
    for n in notes:
        c = wh.cell(row=nr+1, column=1, value=n); c.font = NORMAL; c.alignment = WRAP
        wh.merge_cells(start_row=nr+1, start_column=1, end_row=nr+1, end_column=15)
        wh.row_dimensions[nr+1].height = 30
        nr += 1
    return wh

build_headcount()

# =============================== ASSUMPTIONS ================================ #
wa = wb.create_sheet("Assumptions")
wa.sheet_view.showGridLines = False
wa["A1"] = "Assumptions — non-personnel drivers (yellow cells)"; wa["A1"].font = TITLE
wa["A2"] = "Personnel lives on the Headcount sheet. Everything else is here."; wa["A2"].font = SMALL
wa.column_dimensions["A"].width = 4
wa.column_dimensions["B"].width = 34
wa.column_dimensions["C"].width = 12
hdr_row = 3
wa.cell(row=hdr_row, column=1, value="").fill = SECT
wa.cell(row=hdr_row, column=2, value="Driver").font = BOLD; wa.cell(row=hdr_row, column=2).fill = SECT
wa.cell(row=hdr_row, column=3, value="Unit").font = BOLD; wa.cell(row=hdr_row, column=3).fill = SECT
col = 4
for scen in ["Conservative", "Base", "Ambitious"]:
    wa.cell(row=2, column=col, value=scen).font = H2
    wa.cell(row=2, column=col).alignment = Alignment(horizontal="center")
    wa.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+4)
    for yc in ["Y1", "Y2", "Y3", "Y4", "Y5"]:
        c = wa.cell(row=hdr_row, column=col, value=yc); c.font = BOLD; c.fill = SECT
        c.alignment = Alignment(horizontal="center"); col += 1
    col += 1
for i, (key, label, unit, cons, base, amb, fmt) in enumerate(ASSUMPTION_ROWS):
    row = ASSUMPTION_START_ROW + i
    wa.cell(row=row, column=2, value=label).font = NORMAL
    wa.cell(row=row, column=3, value=unit).font = SMALL
    col = 4
    for scen_vals in (cons, base, amb):
        for v in scen_vals:
            c = wa.cell(row=row, column=col, value=v)
            c.number_format = fmt; c.fill = INPUT; c.alignment = RIGHT; c.font = NORMAL
            col += 1
        col += 1
    wa.row_dimensions[row].height = 16

# =============================== MODEL SHEETS ============================== #
MODEL_TITLE = {
    "Conservative": "Model — Conservative (lean volunteer, DACH-only)",
    "Base":         "Model — Base (steady regional platform, DACH core)",
    "Ambitious":    "Model — Ambitious (pan-European, staffed org)",
}

def build_model(scen):
    wm = wb.create_sheet(f"Model - {scen}")
    wm.sheet_view.showGridLines = False
    wm.column_dimensions["A"].width = 40
    for c in MODEL_COLS: wm.column_dimensions[c].width = 15
    wm["A1"] = MODEL_TITLE[scen]; wm["A1"].font = TITLE
    wm["A2"] = "Personnel from 'Headcount'; all else from 'Assumptions'."; wm["A2"].font = SMALL
    wm.cell(row=4, column=1, value="EUR / year").font = BOLD; wm.cell(row=4, column=1).fill = SECT
    for idx, _ in enumerate(MODEL_COLS):
        c = wm.cell(row=4, column=2+idx, value=f"Year {idx+1}"); c.font = BOLD; c.fill = SECT; c.alignment = Alignment(horizontal="center")

    R = {
        "rev_hdr": 5, "commission": 6, "gm": 7, "events_gross": 8, "lot": 9, "cert": 10, "b2b": 11,
        "grants": 12, "donations": 13, "bridge": 14, "total_rev": 15,
        "dir_hdr": 17, "pay_proc": 18, "event_cogs": 19, "lot_ops": 20, "cert_ops": 21,
        "total_direct": 22, "gross": 23,
        "op_hdr": 25, "infra": 26, "space": 27, "lot_inv": 28, "personnel": 29, "legal": 30, "marketing": 31,
        "total_op": 32, "op_result": 33,
        "tax": 35, "net": 36, "cumulative": 37,
        "memo_hdr": 39, "mission": 40, "commercial": 41, "init_rev": 42, "init_cost": 43,
        "fte": 44, "pers_pct": 45, "op_excl_bridge": 46,
    }
    for h, lab in [("rev_hdr","REVENUE"),("dir_hdr","DIRECT COSTS"),("op_hdr","OPERATING COSTS"),("memo_hdr","MEMO")]:
        cell = wm.cell(row=R[h], column=1, value=lab); cell.font = BOLD; cell.fill = (MEMO if h=="memo_hdr" else SECT)

    labels = {
        "commission":"Commission on paid seats","gm":"GM membership subscriptions","events_gross":"Branded events (gross)",
        "lot":"Library of things (borrowing tier)","cert":"Certification program fees","b2b":"B2B events (commission)",
        "grants":"Grants","donations":"Donations (recurring + one-off)","bridge":"Working-capital bridge (capacity grant / founder loan)","total_rev":"Total revenue",
        "pay_proc":"Payment processing fees","event_cogs":"Event COGS (venue / catering / swag)",
        "lot_ops":"Library shrinkage & replenishment","cert_ops":"Certification ops (assessment + curriculum)",
        "total_direct":"Total direct costs","gross":"Gross contribution",
        "infra":"Infrastructure & hosting","space":"Community HQ (rent + utils + fit-out)",
        "lot_inv":"Library inventory (partnership-sourced)","personnel":"Personnel (from Headcount sheet)",
        "legal":"Legal / admin / accounting (outsourced)","marketing":"Marketing & community",
        "total_op":"Total operating costs","op_result":"Operating result",
        "tax":"Tax provision (illustrative, gemeinnützig)","net":"Net surplus / (deficit)","cumulative":"Cumulative surplus / (deficit)",
        "mission":"Mission-funding reliance (grants+donations)","commercial":"Commercial revenue (all earned streams)",
        "init_rev":"New-initiative revenue (lib+cert+B2B)","init_cost":"New-initiative cost (space+inv+lib+cert ops)",
        "fte":"Total FTE (from Headcount)","pers_pct":"Personnel as % of revenue",
        "op_excl_bridge":"Operating result excl. bridge (underlying business)",
    }
    bold_keys = ("total_rev","total_direct","gross","total_op","op_result","net","cumulative")
    for k, lab in labels.items():
        wm.cell(row=R[k], column=1, value=lab).font = (BOLD if k in bold_keys else NORMAL)

    acols = SCENARIO_COLS[scen]
    for yi, mc in enumerate(MODEL_COLS):
        ac = acols[yi]
        prev = MODEL_COLS[yi-1] if yi > 0 else None
        A = "'Assumptions'!"
        g = lambda key: f"{A}{ac}{ROW_OF[key]}"
        hc = f"'Headcount'!{ac}"   # same column letter maps scenario-year
        wm[f"{mc}{R['commission']}"] = f"={g('paid_seats')}*{g('seat_price')}*{g('comm_rate')}"
        wm[f"{mc}{R['gm']}"] = f"={g('gm_count')}*{g('gm_price')}*12"
        wm[f"{mc}{R['events_gross']}"] = f"={g('events')}*{g('att')}*{g('evt_price')}"
        wm[f"{mc}{R['lot']}"] = f"={g('lot_members')}*{g('lot_price')}*12"
        wm[f"{mc}{R['cert']}"] = f"={g('certs')}*{g('cert_fee')}"
        wm[f"{mc}{R['b2b']}"] = f"={g('b2b_events')}*{g('b2b_att')}*{g('b2b_price')}*{g('b2b_comm')}"
        wm[f"{mc}{R['grants']}"] = f"={g('grants')}"
        wm[f"{mc}{R['donations']}"] = f"={g('donors')}*{g('don_avg')}+{g('don_oneoff')}"
        wm[f"{mc}{R['bridge']}"] = f"={g('bridge')}"
        wm[f"{mc}{R['total_rev']}"] = f"=SUM({mc}{R['commission']}:{mc}{R['bridge']})"
        comm_rev = (f"{mc}{R['commission']}+{mc}{R['gm']}+{mc}{R['events_gross']}"
                    f"+{mc}{R['lot']}+{mc}{R['cert']}+{mc}{R['b2b']}")
        wm[f"{mc}{R['pay_proc']}"] = f"={g('pay_rate')}*({comm_rev})"
        wm[f"{mc}{R['event_cogs']}"] = f"={g('events')}*{g('att')}*{g('evt_cogs')}"
        wm[f"{mc}{R['lot_ops']}"] = f"={g('lot_ops')}"
        wm[f"{mc}{R['cert_ops']}"] = f"={g('cert_dev')}+{g('certs')}*{g('cert_ops')}"
        wm[f"{mc}{R['total_direct']}"] = f"=SUM({mc}{R['pay_proc']}:{mc}{R['cert_ops']})"
        wm[f"{mc}{R['gross']}"] = f"={mc}{R['total_rev']}-{mc}{R['total_direct']}"
        wm[f"{mc}{R['infra']}"] = f"={g('infra')}*12"
        wm[f"{mc}{R['space']}"] = f"={g('locations')}*{g('space_cost')}*12+{g('space_setup')}"
        wm[f"{mc}{R['lot_inv']}"] = f"={g('lot_inv')}"
        wm[f"{mc}{R['personnel']}"] = f"={hc}{HC_COST_ROW}"
        wm[f"{mc}{R['legal']}"] = f"={g('legal')}"
        wm[f"{mc}{R['marketing']}"] = f"={g('marketing')}"
        wm[f"{mc}{R['total_op']}"] = f"=SUM({mc}{R['infra']}:{mc}{R['marketing']})"
        wm[f"{mc}{R['op_result']}"] = f"={mc}{R['gross']}-{mc}{R['total_op']}"
        wm[f"{mc}{R['tax']}"] = (f"=MAX(0,({comm_rev}-{mc}{R['pay_proc']}-{mc}{R['event_cogs']}"
                                 f"-{mc}{R['lot_ops']}-{mc}{R['cert_ops']}-45000))*0.15")
        wm[f"{mc}{R['net']}"] = f"={mc}{R['op_result']}-{mc}{R['tax']}"
        wm[f"{mc}{R['cumulative']}"] = (f"={mc}{R['net']}" if prev is None else f"={prev}{R['cumulative']}+{mc}{R['net']}")
        wm[f"{mc}{R['mission']}"] = f"=({mc}{R['grants']}+{mc}{R['donations']})/{mc}{R['total_rev']}"
        wm[f"{mc}{R['commercial']}"] = f"={comm_rev}"
        wm[f"{mc}{R['init_rev']}"] = f"={mc}{R['lot']}+{mc}{R['cert']}+{mc}{R['b2b']}"
        wm[f"{mc}{R['init_cost']}"] = f"={mc}{R['space']}+{mc}{R['lot_inv']}+{mc}{R['lot_ops']}+{mc}{R['cert_ops']}"
        wm[f"{mc}{R['fte']}"] = f"={hc}{HC_FTE_ROW}"
        wm[f"{mc}{R['pers_pct']}"] = f"={mc}{R['personnel']}/{mc}{R['total_rev']}"
        wm[f"{mc}{R['op_excl_bridge']}"] = f"={mc}{R['op_result']}-{mc}{R['bridge']}"

    eur_rows = [R[k] for k in ["commission","gm","events_gross","lot","cert","b2b","grants","donations","bridge",
                "total_rev","pay_proc","event_cogs","lot_ops","cert_ops","total_direct","gross","infra",
                "space","lot_inv","personnel","legal","marketing","total_op","op_result","tax","net","cumulative",
                "commercial","init_rev","init_cost","op_excl_bridge"]]
    for r_ in eur_rows:
        for mc in MODEL_COLS:
            cell = wm[f"{mc}{r_}"]; cell.number_format = EUR
            if r_ in (R["total_rev"],R["total_direct"],R["gross"],R["total_op"],R["op_result"]): cell.font = BOLD; cell.fill = TOTAL
            if r_ in (R["net"],R["cumulative"]): cell.font = Font(name="Inter",size=11,bold=True,color="065F46"); cell.fill = TOTAL
            if r_ in (R["init_rev"],R["init_cost"],R["op_excl_bridge"]): cell.font = SMALL
    for mc in MODEL_COLS:
        wm[f"{mc}{R['mission']}"].number_format = "0.0%"; wm[f"{mc}{R['mission']}"].font = SMALL
        wm[f"{mc}{R['fte']}"].number_format = "0.00"; wm[f"{mc}{R['fte']}"].font = SMALL
        wm[f"{mc}{R['pers_pct']}"].number_format = "0.0%"; wm[f"{mc}{R['pers_pct']}"].font = SMALL
    return wm

for scen in ["Conservative","Base","Ambitious"]:
    build_model(scen)

# =============================== SUMMARY ==================================== #
wsu = wb.create_sheet("Summary")
wsu.sheet_view.showGridLines = False
wsu.column_dimensions["A"].width = 42
for c in ["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P"]:
    wsu.column_dimensions[c].width = 13
wsu["A1"] = "Summary — three scenarios compared"; wsu["A1"].font = TITLE
wsu["A2"] = "All figures pulled live from the model + headcount sheets."; wsu["A2"].font = SMALL
wsu.cell(row=4, column=1, value="Metric").font = BOLD; wsu.cell(row=4,column=1).fill=SECT
col = 2
for scen in ["Conservative","Base","Ambitious"]:
    wsu.cell(row=3, column=col, value=scen).font = H2
    wsu.cell(row=3, column=col).alignment = Alignment(horizontal="center")
    wsu.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+4)
    for yi in range(5):
        c = wsu.cell(row=4, column=col, value=f"Y{yi+1}"); c.font=BOLD; c.fill=SECT; c.alignment=Alignment(horizontal="center")
        col += 1
    col += 1
# (label, model_row, fmt, bold)
metrics = [
    ("Total revenue",                        15, EUR, True),
    ("Working-capital bridge drawn",         14, EUR, False),
    ("Operating result excl. bridge",        46, EUR, False),
    ("Personnel cost",                       29, EUR, False),
    ("Total FTE",                            44, "0.00", False),
    ("Personnel as % of revenue",            45, "0.0%", False),
    ("New-initiative revenue (lib+cert+B2B)",42, EUR, False),
    ("New-initiative cost",                  43, EUR, False),
    ("Total costs (direct + operating)",     None, EUR, False),
    ("Net surplus / (deficit)",              36, EUR, True),
    ("Cumulative surplus / (deficit)",       37, EUR, True),
    ("Mission-funding reliance",             40, "0.0%", False),
]
r = 5
for label, mrow, fmt, bold in metrics:
    wsu.cell(row=r, column=1, value=label).font = (BOLD if bold else NORMAL)
    col = 2
    for scen in ["Conservative","Base","Ambitious"]:
        sheet = f"'Model - {scen}'"
        for yi, mc in enumerate(MODEL_COLS):
            f = (f"={sheet}!{mc}22+{sheet}!{mc}32" if mrow is None else f"={sheet}!{mc}{mrow}")
            cell = wsu.cell(row=r, column=col, value=f); cell.number_format = fmt
            if bold: cell.font = BOLD
            if mrow == 37: cell.font = Font(name="Inter", size=11, bold=True, color="065F46")
            col += 1
        col += 1
    r += 1
r += 1
wsu.cell(row=r, column=1, value="Reading the model").font = H2; r += 1
notes = [
    "• Conservative stays volunteer-tiny (≤0.65 FTE) and net-positive every year — no bridge needed.",
    "• Base (Commercial Engine path) needs a €170k working-capital bridge drawn Y1–Y4 (50/80/10/30). The "
    "underlying operating business reaches break-even at Y3; with the bridge every year is net-positive and "
    "Y5 produces a genuine €35k surplus.",
    "• Ambitious needs a €115k bridge across Y1–Y2 only; self-sustaining from Y3 and produces a €432k surplus by Y5.",
    "• The Y4 bridge draw (€30k in Base) is larger than Y3's because Y4 carries BOTH the operating dip "
    "(team scales ahead of commercial ramp) AND the first year of corporation tax (commercial surplus crosses "
    "the €45k exemption). The bridge smooths both.",
    "• 'Operating result excl. bridge' (memo line) shows the underlying business honestly: loss-making Y1–Y4, "
    "break-even Y3, surplus Y5. The bridge is the explicit instrument that funds that gap.",
    "• Personnel runs ~50–62% of revenue early, settling to ~40–49% by Y5 — normal for a service/community nonprofit.",
]
for n in notes:
    c = wsu.cell(row=r, column=1, value=n); c.font = NORMAL; c.alignment = WRAP
    wsu.merge_cells(start_row=r, start_column=1, end_row=r, end_column=16); wsu.row_dimensions[r].height = 30; r += 1

# =============================== RESEARCH =================================== #
wr = wb.create_sheet("Research & Benchmarks")
wr.sheet_view.showGridLines = False
wr.column_dimensions["A"].width = 32; wr.column_dimensions["B"].width = 28
wr.column_dimensions["C"].width = 54; wr.column_dimensions["D"].width = 52
wr["A1"] = "Research & Benchmarks — what's reasonable, with sources"; wr["A1"].font = TITLE
wr["A2"] = "Used to ground the assumptions and salary costs. Every row links to its primary source."; wr["A2"].font = SMALL
LINK = Font(name="Inter", size=10, color="1D4ED8", underline="single")
def hyperlink(cell_ref, url, display):
    cell_ref.value = display; cell_ref.hyperlink = url; cell_ref.font = LINK; cell_ref.alignment = WRAP

bench = [
    ("", "Source", "Figure", "Why it matters", "Link"),
    # salaries & staffing (NEW)
    ("Salaries & staffing", "Laravel-Entwickler — Laravel salary guide",
     "Germany: junior €30–45k, mid €50–65k, senior €70–90k (avg ~€55k)",
     "Anchors the Engineering role costs. All-in cost = these gross figures × ~1.22 employer overhead.",
     "https://www.laravel-entwickler.de/en/laravel-developer-salary-a-comprehensive-guide/"),
    ("Salaries & staffing", "Glassdoor — PHP Laravel Developer salary Germany",
     "Avg €67,000 (range €56,500–€70,000)",
     "Cross-check for senior/mid engineering gross salaries in Germany.",
     "https://www.glassdoor.com/Salaries/germany-laravel-php-developer-salary-SRCH_IL.0,7_IN96_KO8,29.htm"),
    ("Salaries & staffing", "GermanTechJobs — Laravel salary Germany",
     "Avg €58,000 / median €57,500",
     "Independent market aggregate for Laravel roles in Germany.",
     "https://germantechjobs.de/en/salaries/Laravel/all/all"),
    ("Salaries & staffing", "PayScale — Program Manager, Non-Profit (Germany)",
     "Avg €49,974 (early €42k, mid €58k)",
     "Anchors Community Manager / Programme Lead costs in the nonprofit sector.",
     "https://www.payscale.com/research/DE/Job=Program_Manager%2C_Non-Profit_Organization/Salary"),
    ("Salaries & staffing", "SalaryExpert — Non-Profit Director (Germany)",
     "Avg €42,646 (entry €31,696)",
     "Reference for fundraising/operations leadership roles; nonprofit-sector ceiling.",
     "https://www.salaryexpert.com/salary/job/non-profit-director/germany"),
    ("Salaries & staffing", "SalaryExplorer — Fundraising & Non-Profit (Germany)",
     "Median €2,660/mo (~€32k), avg €2,810/mo, up to €5,670/mo (~€68k)",
     "Fundraising/Grants Lead and Partnerships Manager salary band (nonprofit field).",
     "http://www.salaryexplorer.com/salary-survey.php?loc=81&loctype=1&job=5&jobtype=1"),
    ("Salaries & staffing", "EngageAnywhere — Employee costs in Germany",
     "Employer overhead ~21–23% (Arbeitgeberanteil)",
     "Drives the gross × 1.22 multiplier used for all-in role costs. Pension 9.3%, health 7.3%, unemployment 1.3%, care 1.525%.",
     "https://engageanywhere.com/blog/calculating-employee-costs-in-germany-step-by-step/"),
    ("Salaries & staffing", "FMC Group — Total cost of employment in Germany",
     "Total employer cost 20–26% above gross",
     "Confirms the all-in multiplier band; notes social-security contribution caps flatten senior costs.",
     "https://fmcgroup.com/employment-cost-germany/"),
    # German donations
    ("German donations", "IW Köln — Fast jeder zweite Deutsche spendet",
     "~€415/donor/yr avg (median ~€300)",
     "German donors give less per head than the global average. Model uses €45–125/yr per recurring donor. "
     "Note: site may return HTTP 401 to bots/HEAD; opens normally in a browser.",
     "https://www.iwkoeln.de/studien/dominik-h-enste-jennifer-potthoff-fast-jeder-zweite-deutsche-spendet.html"),
    ("German donations", "DIW SOEPpapers — Spenden in Deutschland",
     "Median €200→€300 (2009–2017)",
     "Methodology detail behind German donation-volume estimates.",
     "https://www.diw.de/documents/publikationen/73/diw_01.c.738864.de/diw_sp1074.pdf"),
    ("German donations", "World Giving Report (CAF / Maecenata)",
     "46–49% of Germans donate (global 61–64%)",
     "Germany is a low-giving culture; transparency lever for fundraising.",
     "https://www.maecenata.eu/2026/06/03/world-giving-report-2025-deutschland-spendet-weiterhin-unterdurchschnittlich"),
    ("German donations", "Giving in Germany Report 2025 (CAF / Maecenata)",
     "0.39% of income donated; rank 100/101",
     "Companion country report; useful for funder conversations.",
     "https://www.maecenata.eu/2025/07/23/giving-in-germany-report-2025-deutschland-zaehlt-zu-den-schlusslichtern-beim-spenden-weltweit/"),
    ("German donations", "DZI — Spendenstatistik",
     "€5.1bn vs €10.3bn volume estimates",
     "Why different surveys disagree.",
     "https://www.dzi.de/spendenberatung/spendenauskunfte-und-information/spendenstatistik/"),
    # recurring giving
    ("Recurring giving", "Neon One — Recurring Giving Statistics 2026",
     "~$938 / recurring donor / yr",
     "Upper bound for engaged recurring donors (US); Ambitious scenario only.",
     "https://neonone.com/resources/blog/recurring-giving-statistics/"),
    ("Recurring giving", "Qgiv — Fundraising Statistics",
     "Recurring donors avg $287/yr (2022)",
     "Sanity-check band for recurring donation assumptions.",
     "https://www.qgiv.com/blog/fundraising-statistics/"),
    # EU grants
    ("EU grants", "CERV Programme 2026 — guide (global-disruption)",
     "€75k–€500k+; 90% funded; transnational",
     "Civic engagement, inclusion, digital participation. Realistic first win year 2–3 with an NGO partner.",
     "https://global-disruption.com/eu-calls/cerv/index.html"),
    ("EU grants", "EUFundingPortal — CERV calls",
     "Town twinning €8.5k–€50.7k; civic min €75k",
     "Concrete call budgets and per-project ranges.",
     "https://eufundingportal.eu/programme/cerv/"),
    ("EU grants", "EUFundingPortal — CERV Programme overview",
     "Per-call budgets and deadlines",
     "Companion overview with current open calls.",
     "https://eufundingportal.eu/cerv-programme/"),
    ("EU grants", "Creative Europe — calls (EUFundingPortal)",
     "Projects up to €500k",
     "Culture + social-cohesion strand; 'belonging through play' maps to social-resilience objective.",
     "https://eufundingportal.eu/programme/creative-europe/"),
    ("EU grants", "Culture Action Europe — next Creative Europe / AgoraEU",
     "€8.6bn proposed 2028–2034 (€1.8bn culture)",
     "Forward look on EU culture funding.",
     "https://cultureactioneurope.org/news/proposed-e8-6-billion-for-culture-and-democracy-in-the-next-eu-budget/"),
    ("EU grants", "Culture Action Europe — leaked next programme detail",
     "Strand structure + social conditionality",
     "Deeper detail on the next Creative Europe strands.",
     "https://cultureactioneurope.org/news/the-next-creative-europe-programme-has-been-leaked/"),
    ("EU grants", "Erasmus+ — overview (EUFundingPortal)",
     "Inclusion, diversity & digital transformation",
     "Youth / non-formal-education angle; programmatic grants tied to community events.",
     "https://eufundingportal.eu/erasmus-plus/"),
    # German foundations
    ("German foundations", "Aktion Mensch — Förderung",
     "Inclusion project grants (varies)",
     "Major German lottery-funded grantmaker; application via DIAS; nachrangig to public funds. Many calls paused.",
     "https://www.aktion-mensch.de/foerderung"),
    ("German foundations", "reflecta Fördermittelkompass — Social Support Germany",
     "Project grants typically €2k–€50k",
     "Aggregates ESF+, BMFSFJ, BMAS, Aktion Mensch, TV Lottery. Ideal year-1 funding.",
     "https://foerdermittelkompass.reflecta.org/foerderungen/soziales-foerderungen-deutschland?locale=en"),
    ("German foundations", "Bund.de Förderdatenbank",
     "Federal funding search",
     "Primary search portal for all federal/state funding programmes.",
     "https://www.foerderdatenbank.de/"),
    # German tax law
    ("German tax law", "Council on Foundations — Nonprofit Law in Germany",
     "§5(1) Nr.9 KStG; Zweckbetrieb 7% VAT",
     "Purpose-serving revenue exempt from KSt/GewSt; reduced VAT for Zweckbetrieb services.",
     "https://cof.org/content/nonprofit-law-germany"),
    ("German tax law", "GermanCompanyFormation — Non-Profits in Germany",
     "Commercial surplus taxed above €45,000/yr",
     "Steuerpflichtiger wirtschaftlicher Geschäftsbetrieb vs Zweckbetrieb; drives the tax line.",
     "https://germancompanyformation.com/guides/what-is-non-profit-organizations-in-germany"),
    ("German tax law", "Liesegang Partner — Tax-Exempt Status for NPOs",
     "gemeinnützige Zwecke (§52 AO) definition",
     "What qualifies as public-benefit; supports e.V. vs gGmbH choice.",
     "https://www.liesegang-partner.com/knowhow/corporate-law/translate-to-englisch-unlocking-tax-benefits-how-non-profits-in-germany-can-achieve-tax-exempt-status"),
    # library of things
    ("Library of things", "Share Shed — Setting Up & Sustaining a LoT",
     "Membership income < full costs in early years",
     "Lending revenue alone doesn't cover rent/staff early on; grants + volunteers + in-kind space needed. Inventory often donated/discounted.",
     "https://www.shareshed.org.uk/wp-content/uploads/2024/12/Setting-Up-and-Sustaining-a-Library-of-Things-FINAL-2.pdf"),
    ("Library of things", "Shareable — How to start a Library of Things",
     "Subscription + premium-item fees most sustainable",
     "Anchors the library-tier fee model; sliding-scale for affordability.",
     "https://www.shareable.net/how-to/how-to-start-a-library-of-things/"),
    ("Library of things", "Shareable — LoT Toolkit (myTurn)",
     "Bought lending platform ~€2,700/yr (16 locks + SW)",
     "Build-vs-buy reference. Roundup builds its own, so NOT a modelled cost; dev effort sits in Personnel.",
     "https://shareable.net/library-of-things-toolkit"),
    # space & inventory
    ("Space & inventory", "FinancialModelExcel — Indie board game pub startup",
     "Game library $10k–$30k retail; lease deposit $5k–$15k",
     "Retail upper bound that partnership sourcing cuts ~70%; anchors HQ fit-out.",
     "https://financialmodelexcel.com/blogs/cost-open/indie-board-game-pub"),
    ("Space & inventory", "StartupFinancialProjection — Board game cafe",
     "Library 500 games $10k–$20k; rent $5k–$25k deposit",
     "Sanity-checks inventory and rent ranges; roundup HQ is leaner than a full cafe.",
     "https://startupfinancialprojection.com/blogs/capex/board-game-cafe"),
    # B2B events
    ("B2B events", "TeamBonding — Team building cost",
     "In-person from $3,500; $45–$60/person avg",
     "Anchors B2B per-attendee pricing (model uses €35–€65, European lower end).",
     "https://www.teambonding.com/cost/"),
    ("B2B events", "Leaders Institute — Corporate team building cost",
     "$35–$75/person typical",
     "Confirms the per-person band for professional facilitated events.",
     "https://www.leadersinstitute.com/get-a-price/"),
    ("B2B events", "IRL Game Shop — Corporate team building",
     "$20–$25/person board-game events",
     "Tabletop-specific corporate pricing; supports the lower end of B2B price assumption.",
     "https://www.irlgameshop.com/service/corporate-team-building-events/"),
    # tabletop grants
    ("Tabletop grants", "TTGDA — Scholarships list",
     "$250–$2,000 micro-grants (US, niche)",
     "Niche gives small grants; useful for community micro-programs, not core ops.",
     "https://www.ttgda.org/scholarships"),
    ("Tabletop grants", "Washington State Library — TTRPG Innovation Grants",
     "Up to $2,000 (libraries)",
     "State-scale hobby funding; model for a library-partnership programme.",
     "https://www.sos.wa.gov/about-office/news/2025/libraries-washington-apply-now-tabletop-gaming-material-grants"),
    ("Tabletop grants", "ALA Game On! Grants 2025",
     "$2,000 (doubled by CAH donation)",
     "Library-channel example; shows corporate co-funding pattern — a precedent for publisher donations.",
     "https://games.ala.org/game-on-grant-applications-open-for-2025/"),
    ("Tabletop grants", "Gen Con — Participation Grants",
     "$550 stipend + 4-day badge",
     "Convention attendance grant; model for roundup.events accessibility grants.",
     "https://www.gencon.com/gen-con-indy/participation-grants"),
    # event economics
    ("Event economics", "Dataintelo — Fan Conventions Market Report 2034",
     "Tickets 31% / merch 27% / sponsorship 21% / F&B 10%",
     "Anchors branded-events margin assumption (~20–40% net) for small events.",
     "https://dataintelo.com/report/fan-conventions-market"),
]

row = 4; last_group = None
for entry in bench:
    group, src, fig, why, url = entry
    if src == "Source":
        for col, val in enumerate([src, fig, why, url], start=1):
            c = wr.cell(row=row, column=col, value=val); c.font = BOLD; c.fill = SECT; c.alignment = WRAP
        wr.row_dimensions[row].height = 18; row += 1; continue
    if group != last_group:
        gc = wr.cell(row=row, column=1, value=group); gc.font = H2; gc.fill = GROUP_FILL
        wr.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        wr.row_dimensions[row].height = 18; last_group = group; row += 1
    wr.cell(row=row, column=1, value=src).font = NORMAL; wr.cell(row=row, column=1).alignment = WRAP
    wr.cell(row=row, column=2, value=fig).font = NORMAL; wr.cell(row=row, column=2).alignment = WRAP
    wr.cell(row=row, column=3, value=why).font = NORMAL; wr.cell(row=row, column=3).alignment = WRAP
    hyperlink(wr.cell(row=row, column=4), url, url)
    wr.row_dimensions[row].height = 48; row += 1

out = "docs/funding/roundup-games-financial-model.xlsx"
wb.save(out)
print(f"Wrote {out}")
